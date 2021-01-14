import discord
from discord.ext import commands
import openpyxl

xlsx = openpyxl.load_workbook("riscord/members.xlsx")
sheet = xlsx.active

class Verification(commands.Cog):
	def __init__(self, client):
		self.client = client

	@commands.command(aliases=["v"])
	async def verify(self, ctx, student_id):
		for i in range(2, sheet.max_row+1):
			id_val = str(sheet.cell(row=i, column=9).value)
			valid_chars = ["s", "S"]
					
			if id_val[0] not in valid_chars:
				id_val = "s"+id_val
			
			if id_val[0] == valid_chars[1]:
				id_val = "s" + id_val[1:-1]

			if id_val == student_id:
				member = ctx.message.author
				print(member.roles)
				member_role = discord.utils.get(member.guild.roles, name="Member")
				try:
					await member.add_roles(member_role)
				except	:
					pass
				await ctx.send("Welcome to the club :tada:")
				return True

		await ctx.send("User not found")
		return False

def setup(client):
	client.add_cog(Verification(client)) 