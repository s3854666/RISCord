import openpyxl

async def verify_user(student_id):
	xlsx = openpyxl.load_workbook('members.xlsx')
	sheet = xlsx.active

	for i in range(2, sheet.max_row+1):
		id_val = str(sheet.cell(row=i, column=9).value)
		valid_chars = ["s", "S"]
				
		if id_val[0] not in valid_chars:
			id_val = "s"+id_val
		
		if id_val[0] == valid_chars[1]:
			id_val = "s" + id_val[1:-1]

		if id_val == student_id:
			return True
		
	return False
